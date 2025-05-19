from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from parse_vigiaccess_html import Drug
from datetime import datetime
import csv
import os

# Возрастные группы
AGE_GROUPS = [
    "0 - 27 days",
    "28 days to 23 months",
    "2 - 11 years",
    "12 - 17 years",
    "18 - 44 years",
    "45 - 64 years",
    "65 - 74 years",
    "≥ 75 years",
    "Unknown"
]

# Список полов
SEX_GROUPS = [
    "Female",
    "Male",
    "Unknown"
]

# Константы
ROW_NAME_RU = 1
ROW_NAME_EN = 2
ROW_REPORT_COUNT = 3
ROW_AGE_START = 5
ROW_SEX_START = ROW_AGE_START + len(AGE_GROUPS) + 1
ROW_EFFECTS_START = ROW_SEX_START + len(SEX_GROUPS)

def set_bold_text(sheet, row, column, text):
    cell = f"{get_column_letter(column)}{row}"
    sheet[cell] = text
    sheet[cell].font = sheet[cell].font.copy(bold=True)

def fill_cell(sheet, row, column, color = "CCFFCC"):
    fill_color = PatternFill(start_color=color, end_color=color, fill_type="solid")
    sheet.cell(row=row, column=column).fill = fill_color

def get_side_e_set(drug_list):
    # Словарь для хранения уникальных подпобочных элементов для каждого побочного эффекта
    side_e_set = {}

    # Для каждого побочного эффекта уникальные подпобочные элементы
    for drug in drug_list:
        if drug.group_side_e:
            for group_side_e in drug.group_side_e:
                if group_side_e not in side_e_set:
                    side_e_set[group_side_e] = set()
                for side_e in drug.group_side_e[group_side_e]["side_e_dict"]:
                    side_e_set[group_side_e].add(side_e)

    return side_e_set

# Основная функция
def save_to_excel(drug_list, side_e_set, output_filename):
    wb = Workbook()
    sheet = wb.active

    # Заголовки
    sheet.cell(row=ROW_NAME_RU, column=1, value="Название ru")
    sheet.cell(row=ROW_NAME_EN, column=1, value="Название en")
    sheet.cell(row=ROW_REPORT_COUNT, column=1, value="Количество обращений")

    # Возраст
    fill_cell(sheet, ROW_AGE_START - 1, 1)
    set_bold_text(sheet, ROW_AGE_START - 1, 1, "Возраст")
    for idx, age_name in enumerate(AGE_GROUPS):
        sheet.cell(row=ROW_AGE_START + idx, column=1, value=age_name)

    # Пол
    fill_cell(sheet, ROW_SEX_START - 1, 1)
    set_bold_text(sheet, ROW_SEX_START - 1, 1, "Пол")
    for idx, sex_name in enumerate(SEX_GROUPS):
        sheet.cell(row=ROW_SEX_START + idx, column=1, value=sex_name)

    # Побочные эффекты
    effect_row = ROW_EFFECTS_START
    for group in sorted(side_e_set):
        fill_cell(sheet, effect_row, 1)
        set_bold_text(sheet, effect_row, 1, group)
        
        for i in range(len(drug_list)):
            fill_cell(sheet, effect_row, i+2)
            set_bold_text(sheet, effect_row, i+2, 0)

        effect_row += 1
        for sub_effect in sorted(side_e_set[group]):
            sheet.cell(row=effect_row, column=1, value=sub_effect)

            for i in range(len(drug_list)):
                sheet.cell(row=effect_row, column=i+2, value=0)
            effect_row += 1

    # Препараты
    for i, drg in enumerate(drug_list):
        col = i + 2

        # Метаданные
        sheet.cell(row=ROW_NAME_RU, column=col, value=drg.name_ru)
        sheet.cell(row=ROW_NAME_EN, column=col, value=drg.name_en)
        sheet.cell(row=ROW_REPORT_COUNT, column=col, value=drg.report_count)

        if drg.name_ru == "Риамиловир":
            print("drg.report_count:", drg.report_count)
        
        if not drg.group_side_e:
            continue

        # Возраст
        fill_cell(sheet, ROW_AGE_START - 1, col)
        for idx, age_name in enumerate(AGE_GROUPS):
            sheet.cell(row=ROW_AGE_START + idx, column=col, value=drg.age_info.get(age_name, 0))

        # Пол
        fill_cell(sheet, ROW_SEX_START - 1, col)
        for idx, sex_name in enumerate(SEX_GROUPS):
            sheet.cell(row=ROW_SEX_START + idx, column=col, value=drg.sex_info.get(sex_name, 0))

        # Побочные эффекты
        current_row = ROW_EFFECTS_START
        current_group = None

        while sheet.cell(row=current_row, column=1).value is not None:
            cell_value = sheet.cell(row=current_row, column=1).value

            if cell_value in drg.group_side_e:
                current_group = cell_value
                fill_cell(sheet, current_row, col)
                set_bold_text(sheet, current_row, col, drg.group_side_e[current_group].get("cases", 0))
            elif current_group:
                value = (
                    drg.group_side_e.get(current_group, {})
                    .get("side_e_dict", {})
                    .get(cell_value, 0)
                )
                sheet.cell(row=current_row, column=col, value=value)

            current_row += 1

    wb.save(output_filename)

if __name__ == "__main__":
    DIR_SOURCE_HTML  = "parse_vigiaccess\\data\\vigiaccess_html"
    LINKS_SIDE_EFFECT_FILENAME = "make_side_effect_dataset\\data\\drugs_table4.csv"


    # Читаем CSV-файл
    with open(LINKS_SIDE_EFFECT_FILENAME, newline='', encoding='utf-8') as csvfile:
        drug_rows = list(csv.DictReader(csvfile, delimiter=';'))  # Читаем сразу весь файл
        total_rows = len(drug_rows)

    drug_list = []
    for drug_row in drug_rows: 

        drug_name_ru = drug_row.get('drug_name_ru', '').strip()
        drug_name_en = drug_row.get('drug_name_en', '').strip()

        file_path = f"{DIR_SOURCE_HTML}\\all_html_{drug_name_en}.html"
        drug_obj = Drug(drug_name_ru, drug_name_en)
        if os.path.exists(file_path):
            with open(file_path, 'r', encoding='utf-8') as file:
                html_content = file.read()
                
            
            drug_obj.parse_vigiaccess_html(html_content)

            drug_list.append(drug_obj)
        else:
            print(f"⚠️ Файл не найден: {file_path}")
            drug_list.append(drug_obj)

    # Поиск уникальных побочек
    side_e_set = get_side_e_set(drug_list)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    save_to_excel(drug_list, side_e_set, f"parse_vigiaccess\\data\\Vigiaccess_{timestamp}.xlsx")

